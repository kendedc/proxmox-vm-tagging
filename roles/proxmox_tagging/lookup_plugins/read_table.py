"""Lookup plugin that reads a csv/tsv/xlsx sheet into a list of row dicts."""

from __future__ import annotations

DOCUMENTATION = r"""
name: read_table
short_description: Read a csv, tsv or xlsx table into a list of dictionaries
description:
  - Reads a tabular file and returns one dict per data row, keyed by the header
    row values.
  - CSV and TSV are handled with the Python standard library, so no extra
    dependency is required. XLSX additionally needs the openpyxl package in the
    execution environment.
  - The format is chosen from the file extension unless O(format) is set.
options:
  _terms:
    description: Path to the table file
    required: true
    type: list
    elements: path
  format:
    description: Force a format instead of detecting it from the extension
    type: str
    choices: [auto, csv, tsv, xlsx]
    default: auto
  sheet:
    description: Worksheet name for xlsx; defaults to the active sheet
    type: str
    default: null
  header_row:
    description: 1-indexed row holding the column headers
    type: int
    default: 1
  delimiter:
    description: Field delimiter override for csv/tsv
    type: str
    default: null
  encoding:
    description: Text encoding; utf-8-sig transparently strips an Excel BOM
    type: str
    default: utf-8-sig
  skip_empty:
    description: Drop rows where every cell is empty
    type: bool
    default: true
"""

EXAMPLES = r"""
- name: Load the tag source table
  ansible.builtin.set_fact:
    rows: "{{ lookup('read_table', '/data/vm_tags.csv') }}"

- name: Load a specific worksheet from a workbook
  ansible.builtin.set_fact:
    rows: "{{ lookup('read_table', '/data/vm_tags.xlsx', sheet='VMs') }}"
"""

RETURN = r"""
_raw:
  description: One dictionary per data row
  type: list
  elements: dict
"""

import csv  # noqa: E402
import datetime  # noqa: E402
import io  # noqa: E402
import os  # noqa: E402

from ansible.errors import AnsibleError  # noqa: E402
from ansible.module_utils.common.text.converters import to_native  # noqa: E402
from ansible.plugins.lookup import LookupBase  # noqa: E402

_EXTENSIONS = {
    ".csv": "csv",
    ".tsv": "tsv",
    ".tab": "tsv",
    ".xlsx": "xlsx",
    ".xlsm": "xlsx",
}

_DELIMITERS = {"csv": ",", "tsv": "\t"}


def _cell_to_text(value):
    """Normalize one cell into a trimmed string"""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (datetime.datetime, datetime.date)):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _build_headers(header_cells):
    """Turn the header row into stable, non-empty column keys"""
    headers = []
    for index, cell in enumerate(header_cells):
        name = _cell_to_text(cell)
        headers.append(name if name else "column_{0}".format(index + 1))
    return tuple(headers)


def _rows_to_records(raw_rows, header_row, skip_empty, source):
    """Turn an iterable of cell tuples into header-keyed dicts"""
    headers = None
    records = []
    for row_number, row in enumerate(raw_rows, start=1):
        if row_number < header_row:
            continue
        if row_number == header_row:
            headers = _build_headers(row)
            continue

        values = [_cell_to_text(cell) for cell in row]
        if skip_empty and not any(values):
            continue

        record = dict(zip(headers, values))
        record["_row"] = row_number
        records.append(record)

    if headers is None:
        raise AnsibleError(
            "header row {0} not found in {1}".format(header_row, source)
        )
    return records


class LookupModule(LookupBase):
    def run(self, terms, variables=None, **kwargs):
        self.set_options(var_options=variables, direct=kwargs)

        results = []
        for term in terms:
            path = self.find_file_in_search_path(variables, "files", term) or term
            fmt = self._resolve_format(path)
            if fmt == "xlsx":
                results.extend(self._read_xlsx(path))
            else:
                results.extend(self._read_delimited(path, fmt))
        return results

    def _resolve_format(self, path):
        """Pick the reader from the option or the file extension"""
        fmt = self.get_option("format") or "auto"
        if fmt != "auto":
            return fmt

        extension = os.path.splitext(path)[1].lower()
        if extension not in _EXTENSIONS:
            raise AnsibleError(
                "cannot infer a format for {0}; pass format=csv, tsv or "
                "xlsx explicitly".format(path)
            )
        return _EXTENSIONS[extension]

    def _read_delimited(self, path, fmt):
        """Read a csv or tsv file with the standard library"""
        delimiter = self.get_option("delimiter") or _DELIMITERS[fmt]
        encoding = self.get_option("encoding") or "utf-8-sig"

        try:
            with io.open(path, "r", encoding=encoding, newline="") as handle:
                raw_rows = list(csv.reader(handle, delimiter=delimiter))
        except UnicodeDecodeError as exc:
            raise AnsibleError(
                "could not decode {0} as {1}; if the sheet was exported from "
                "Excel as 'CSV' rather than 'CSV UTF-8', re-export it or set "
                "encoding=cp1252: {2}".format(path, encoding, to_native(exc))
            )
        except (IOError, OSError) as exc:
            raise AnsibleError(
                "could not read {0}: {1}".format(path, to_native(exc))
            )

        return _rows_to_records(
            raw_rows, self.get_option("header_row"), self.get_option("skip_empty"), path
        )

    def _read_xlsx(self, path):
        """Read one worksheet from a workbook via openpyxl"""
        try:
            import openpyxl
        except ImportError:
            raise AnsibleError(
                "reading {0} requires openpyxl, which is not in this execution "
                "environment; either build an EE with openpyxl or export the "
                "sheet to CSV, which needs no extra dependency".format(path)
            )

        try:
            workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception as exc:
            raise AnsibleError(
                "could not open workbook {0}: {1}".format(path, to_native(exc))
            )

        try:
            sheet_name = self.get_option("sheet")
            if sheet_name:
                if sheet_name not in workbook.sheetnames:
                    raise AnsibleError(
                        "worksheet '{0}' not found in {1}; available: {2}".format(
                            sheet_name, path, ", ".join(workbook.sheetnames)
                        )
                    )
                worksheet = workbook[sheet_name]
            else:
                worksheet = workbook.active

            return _rows_to_records(
                worksheet.iter_rows(values_only=True),
                self.get_option("header_row"),
                self.get_option("skip_empty"),
                path,
            )
        finally:
            workbook.close()
